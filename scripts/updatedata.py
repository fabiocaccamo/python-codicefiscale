# -*- coding: utf-8 -*-

from benedict import benedict
from openpyxl import load_workbook
from slugify import slugify

import fsutil


def _update_countries_data():
    # https://www.anagrafenazionale.interno.it/il-progetto/strumenti-di-lavoro/tabelle-decodifica/
    data_url = "https://www.anagrafenazionale.interno.it/wp-content/uploads/2021/03/tabella_2_statiesteri.xlsx"
    data_path = fsutil.download_file(data_url, __file__, filename="countries.xlsx")

    workbook = load_workbook(filename=data_path, read_only=True)
    sheet = workbook.active

    items = []
    keys = []
    for row in sheet.iter_rows(min_row=1, max_row=1):
        keys = [cell.value for cell in row]
    for row in sheet.iter_rows(min_row=2):
        values = list([cell.value for cell in row])
        items.append(dict(zip(keys, values)))

    workbook.close()
    fsutil.remove_file(data_path)

    data = benedict({"values": items})
    data.standardize()
    # print(data.dump())

    def map_item(item):
        if not item:
            return None
        code = item.get_str("codat").upper()
        if not code:
            return None
        assert len(code) == 4, f"Invalid code: '{code}'"

        name = item.get_str("denominazione").title()
        assert name != "", f"Invalid name: '{name}'"
        name_alt = item.get_str("denominazioneistat").title()
        name_alt_en = item.get_str("denominazioneistat_en").title()
        name_slugs = sorted(
            set(
                filter(
                    bool,
                    [
                        slugify(name),
                        slugify(name_alt),
                        slugify(name_alt_en),
                    ],
                )
            )
        )
        province = "EE"

        date_created = item.get_str("datainiziovalidita")
        date_deleted = item.get_str("datafinevalidita")
        if "9999" in date_deleted:
            date_deleted = ""

        return {
            "active": False if date_deleted else True,
            "code": code,
            "date_created": date_created,
            "date_deleted": date_deleted,
            "name": name,
            "name_alt": name_alt,
            "name_alt_en": name_alt_en,
            "name_slugs": name_slugs,
            "province": province,
        }

    output_data = list(
        filter(bool, [map_item(benedict(item)) for item in data["values"]])
    )
    output_data = sorted(output_data, key=lambda item: item["name"])
    output_path = "../codicefiscale/data/countries.json"
    output_abspath = fsutil.join_path(__file__, output_path)
    fsutil.write_file_json(output_abspath, output_data, indent=4, sort_keys=True)


def _update_municipalities_data():
    # https://www.anagrafenazionale.interno.it/il-progetto/strumenti-di-lavoro/tabelle-decodifica/
    data_url = "https://www.anagrafenazionale.interno.it/wp-content/uploads/ANPR_archivio_comuni.csv"
    data = benedict.from_csv(data_url)
    data.standardize()

    def map_item(item):

        status = item.get("stato", "").upper()
        assert len(status) == 1 and status in ["A", "C"], f"Invalid status: '{status}'"
        active = status == "A"

        code = item.get_str("codcatastale").upper()
        assert code == "ND" or len(code) == 4, f"Invalid code: '{code}'"

        name = item.get_str("denominazione_it").title()
        assert name != "", f"Invalid name: {name}"

        name_trans = item.get_str("denomtraslitterata").title()
        name_alt = item.get_str("altradenominazione").title()
        name_alt_trans = item.get_str("altradenomtraslitterata").title()
        name_slugs = sorted(
            set(
                filter(
                    bool,
                    [
                        slugify(name),
                        slugify(name_trans),
                        slugify(name_alt),
                        slugify(name_alt_trans),
                    ],
                )
            )
        )
        province = item.get("siglaprovincia", "").upper()
        assert len(province) == 2, f"Invalid province: '{province}'"

        date_created = item.get_str("dataistituzione")
        date_deleted = item.get_str("datacessazione")
        if "9999" in date_deleted:
            date_deleted = ""

        return {
            "active": active,
            "code": code,
            "date_created": date_created,
            "date_deleted": date_deleted,
            "name": name,
            "name_trans": name_trans,
            "name_alt": name_alt,
            "name_alt_trans": name_alt_trans,
            "name_slugs": name_slugs,
            "province": province,
        }

    output_data = list(
        filter(bool, [map_item(benedict(item)) for item in data["values"]])
    )
    output_data = sorted(output_data, key=lambda item: item["name"])
    output_path = "../codicefiscale/data/municipalities.json"
    output_abspath = fsutil.join_path(__file__, output_path)
    fsutil.write_file_json(output_abspath, output_data, indent=4, sort_keys=True)


def main():
    _update_countries_data()
    _update_municipalities_data()


if __name__ == "__main__":
    main()
